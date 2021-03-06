"""

There are many functions in this module.

"""
from matplotlib import pyplot
import scipy.sparse.linalg as sla
import kwant
import numpy as np
import scipy
import openpyxl


def computeEigenvalues(system):
    sparseMatrix = system.hamiltonian_submatrix(sparse = True)
    eigenvalues = sla.eigs(sparseMatrix, 200)[0]
    print(eigenvalues.real)


def plotConductance(system, energyMaxmum):

    energyMaxmum = energyMaxmum
    r = energyMaxmum / 0.01
    energies = [-1  +  n * 0.01 for n in range(int(r) + 200)]

    data = []

    for energy in energies:
        smatrix = kwant.solvers.default.smatrix(system, energy)
        data.append(smatrix.transmission(0, 1))

    pyplot.figure()
    pyplot.plot(energies, data)
    pyplot.title("Quantum Conductance")
    pyplot.xlabel("energy [t]")
    pyplot.ylabel("conductance [e^2/h]")
    pyplot.show()



def getConductance(system, energyMaxmum, workbook):

    r = energyMaxmum / 0.01
    energies =[-1 + n * 0.01 for n in range(int(r) + 200)]
    data = []

    for energy in energies:
        smatrix = kwant.solvers.default.smatrix(system, energy)
        data.append(smatrix.transmission(0, 1))

    conductanceSheet = workbook.create_sheet("Conductance")
    conductanceSheet.append(energies)
    conductanceSheet.append(data)

    print("OK")


def sortedEigens(ev):
    eigenvalues, eigenvectors = ev
    eigenvalues, eigenvectors = map(np.array, zip(*sorted(zip(eigenvalues, eigenvectors.transpose()))))
    return eigenvalues, eigenvectors.transpose()


def plotWavefunction(system, K):

    hamiltonian_matrix = system.hamiltonian_submatrix(sparse = True)
    eigenvalues, eigenvectors = sortedEigens(sla.eigsh(hamiltonian_matrix.tocsc(), K, sigma = 0))
    a = 0
    for k in range(K):
        a = a + np.abs(eigenvectors[:,k])**2

    kwant.plotter.map(system, a ,colorbar = False, oversampling = 3)


def plotBandstructure(flead, momenta, energyMaxmum):

    bands = kwant.physics.Bands(flead)
    energies = [bands(k) for k in momenta]
    ###

    ###
    pyplot.figure()
    pyplot.plot(momenta, energies)
    pyplot.ylim(-0.005, energyMaxmum)
    pyplot.xlabel("momentum [(lattice constant)^-1]")
    pyplot.ylabel("energy [t]")
    pyplot.show()
    print(bands(0))


def getBandStructure(flead, momenta, length, workbook):

    bands = kwant.physics.Bands(flead)

    band = []
    for subscription in range(length):
        band.append([bands(k)[subscription] for k in momenta])

    bandSheet = workbook.create_sheet("Bands")
    bandSheet.append(momenta)
    for i in range(length):
        bandSheet.append(band[i])

    print("OK")





def plotDensityOfStateMethod(lableToData, energyMaxmum):

    pyplot.figure(figsize=(5, 4))
    x, y = lableToData
    pyplot.plot(x, y)
    pyplot.xlim(0, energyMaxmum)
    pyplot.xlabel("energy [t]")
    pyplot.ylabel("DoS [a.u.]")
    pyplot.show()


def plotDOS(finalizeSystem, energyMaxmum):

    spectrum = kwant.kpm.SpectralDensity(finalizeSystem)
    energies, densities = spectrum()

    plotDensityOfStateMethod((energies, densities), energyMaxmum)


def getDOS(finalizeSystem, workbook):

    spectrum = kwant.kpm.SpectralDensity(finalizeSystem)
    energies, densities = spectrum()
    densities = np.real(densities)

    dosSheet = workbook.create_sheet("DOS")
    dosSheet.append(energy for energy in energies)
    dosSheet.append(density for density in densities)

    print("OK")



def BandEnergy(flead, momentum, length):

    momenta = [-momentum + 0.02 * i for i in range(int(momentum* 100) + 1)]
    bands = kwant.physics.Bands(flead)
    bandEnergy = []
    for subscription in range(length):
        bandEnergy.append([bands(k)[subscription] for k in momenta])

    #bandEnergy1 = [bands(k)[1] for  k in momenta]


    return bandEnergy
