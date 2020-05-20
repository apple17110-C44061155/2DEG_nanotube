import twodegmodel as tm
import kwant
import calculusMathod as cm
import openpyxl


def main(length = 10, width = 10, energyMaxmum = 10, name = "length=10, width=10 system.xlsx"):

    workbook = openpyxl.Workbook()


    system = tm.makeSystem(length, width)
    lead = tm. makeLead(length, width)

    system.attach_lead(lead)
    system.attach_lead(lead.reversed())

    system = system.finalized()

    kwant.plot(system)


    lead = lead.finalized()
    momentum = 6.28
    momenta = [-momentum + 0.02 * i for i in range(int(momentum*100) + 1)]
    cm.plotBandstructure(lead, momenta, energyMaxmum = 10)

    """
    energies, densities = cm.getDOS(system)
    DosSheet = workbook.create_sheet("DOS")
    DosSheet.append(energy for energy in energies)
    DosSheet.append(densitiy for densitiy in densities)
    print("OK")
    """
    """
    momentum = 6.28
    momenta, band = cm.getBandStructure(lead.finalized(), momentum, length)
    BandSheet = workbook.create_sheet()
    BandSheet.append(momenta)
    for i in range(length):
        BandSheet.append(band[i])
    print("OK")
    """
    """
    energies, data = cm.getConductance(system, energyMaxmum)
    ConductanceSheet = workbook.create_sheet("Conductance")
    ConductanceSheet.append(energies)
    ConductanceSheet.append(data)
    print("OK")

    workbook.save(name)
    """
    #cm.computeEigenvalues(system)

if __name__ == '__main__':
    main()
